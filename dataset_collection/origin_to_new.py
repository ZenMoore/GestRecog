from openpyxl.workbook import Workbook

import numpy as np

outwb = Workbook()
wo = outwb.active

careerSheet = outwb.create_sheet('sheet1',0)   #创建的sheet

a=np.ones((6,256))
for colnumber in range(1,7):
    for rownumber in range(1,257):
        careerSheet.cell(row=rownumber,column=colnumber).value =a[colnumber-1][rownumber-1]


outwb.save("sample.xlsx")